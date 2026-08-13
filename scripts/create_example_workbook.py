"""Generate the polished synthetic example component database."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = ["Part_ID","Type","Brand","Model","Variant","Condition","Seller","URL","Date_Found","Price_GBP","Shipping_GBP","Effective_Price_GBP","Warranty_Months","AI_Score","Animation_Score","Gaming_Score","Performance_Score","Reliability_Score","Value_Score","VRAM_GB","RAM_GB","RAM_Type","RAM_Speed_MHz","CPU_Cores","CPU_Threads","CPU_Socket","Motherboard_Socket","Motherboard_RAM_Type","Power_W","PSU_W","PSU_Efficiency","GPU_Length_mm","Case_Max_GPU_Length_mm","Storage_GB","Storage_Type","Cooler_Height_mm","Case_Max_Cooler_Height_mm","Notes"]

def row(pid, kind, brand, model, price, **values):
    base = {"Part_ID":pid,"Type":kind,"Brand":brand,"Model":model,"Variant":"Example","Condition":"New","Seller":"Synthetic example","Price_GBP":price,"Shipping_GBP":0,"Warranty_Months":24,"AI_Score":50,"Animation_Score":50,"Gaming_Score":50,"Performance_Score":50,"Reliability_Score":80,"Value_Score":70,"Notes":"Synthetic user-defined scores; not measured benchmarks."}
    base.update(values); return base

ROWS = [
 row("CPU-001","CPU","AMD","Ryzen 7 Example",230,CPU_Socket="AM5",CPU_Cores=8,CPU_Threads=16,Power_W=105,AI_Score=78,Animation_Score=84,Gaming_Score=82),
 row("CPU-002","CPU","AMD","Ryzen 9 Example",330,CPU_Socket="AM5",CPU_Cores=12,CPU_Threads=24,Power_W=120,AI_Score=90,Animation_Score=94,Gaming_Score=87,Condition="Used",Warranty_Months=6),
 row("GPU-001","GPU","NVIDIA","24GB AI Example",445,Condition="Used",Warranty_Months=3,VRAM_GB=24,GPU_Length_mm=313,Power_W=350,AI_Score=97,Animation_Score=91,Gaming_Score=86),
 row("GPU-002","GPU","NVIDIA","12GB Gaming Example",480,VRAM_GB=12,GPU_Length_mm=270,Power_W=200,AI_Score=78,Animation_Score=82,Gaming_Score=94),
 row("MB-001","Motherboard","Example","B650 Board",130,Motherboard_Socket="AM5",Motherboard_RAM_Type="DDR5",Power_W=45),
 row("RAM-001","RAM","Example","32GB DDR5 Kit",75,RAM_GB=32,RAM_Type="DDR5",RAM_Speed_MHz=6000,Power_W=10,AI_Score=75,Animation_Score=75),
 row("RAM-002","RAM","Example","64GB DDR5 Kit",145,RAM_GB=64,RAM_Type="DDR5",RAM_Speed_MHz=5600,Power_W=14,AI_Score=90,Animation_Score=88),
 row("SSD-001","SSD","Example","1TB NVMe",65,Storage_GB=1000,Storage_Type="NVMe",Power_W=6),
 row("SSD-002","SSD","Example","2TB NVMe",95,Storage_GB=2000,Storage_Type="NVMe",Power_W=7,AI_Score=60,Animation_Score=65),
 row("PSU-001","PSU","Example","850W Gold",90,PSU_W=850,PSU_Efficiency="80+ Gold"),
 row("CASE-001","Case","Example","Airflow Case",70,Case_Max_GPU_Length_mm=350,Case_Max_Cooler_Height_mm=170),
 row("COOL-001","CPU Cooler","Example","Dual Tower Cooler",38,Cooler_Height_mm=157,Power_W=5),
]

def create(path: Path):
    wb=Workbook(); info=wb.active; info.title="README"
    info.append(["PC Parts database – instructions"]); info["A1"].font=Font(size=16,bold=True)
    instructions=["Enter one purchasable option per Parts row; new/used variants are separate rows.","Blue cells are intended inputs. Effective_Price_GBP is a formula (price + shipping).","Scores are user-defined normalised 0–100 values, not real benchmark measurements.","Required compatibility data depends on Type. Run `python -m pc_optimizer validate` before optimisation.","Do not rename headers. Blank non-applicable cells are expected."]
    for text in instructions: info.append([text])
    info.column_dimensions["A"].width=110
    ws=wb.create_sheet("Parts"); ws.append(COLUMNS)
    for item in ROWS:
        ws.append([item.get(c) for c in COLUMNS]); r=ws.max_row; ws.cell(r, COLUMNS.index("Effective_Price_GBP")+1, f"=J{r}+K{r}")
    table=Table(displayName="PCParts",ref=f"A1:AL{ws.max_row}"); table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True); ws.add_table(table)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for c in range(1,len(COLUMNS)+1): ws.column_dimensions[ws.cell(1,c).column_letter].width=min(28,max(12,len(COLUMNS[c-1])+2))
    input_fill=PatternFill("solid",fgColor="DDEBF7")
    formula_col=COLUMNS.index("Effective_Price_GBP")+1
    for cells in ws.iter_rows(min_row=2,max_row=250,max_col=len(COLUMNS)):
        for cell in cells:
            if cell.column != formula_col: cell.fill=input_fill
    kinds='"CPU,GPU,Motherboard,RAM,SSD,PSU,Case,CPU Cooler"'; conditions='"New,Open Box,Refurbished,Used"'
    for formula, col in ((kinds,2),(conditions,6)):
        dv=DataValidation(type="list",formula1=formula); ws.add_data_validation(dv); dv.add(f"{ws.cell(2,col).coordinate}:{ws.cell(250,col).coordinate}")
    for col in range(14,20): ws.conditional_formatting.add(f"{ws.cell(2,col).coordinate}:{ws.cell(250,col).coordinate}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
    path.parent.mkdir(parents=True,exist_ok=True); wb.save(path)

if __name__ == "__main__":
    create(Path("data/example_pc_parts.xlsx")); create(Path("data/pc_parts.xlsx"))
